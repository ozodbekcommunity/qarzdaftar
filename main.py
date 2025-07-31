from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from models import *
from datetime import datetime, date
import os
from dotenv import load_dotenv
from decimal import Decimal
from werkzeug.utils import secure_filename
from peewee import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from peewee import fn

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

# Ma'lumotlar bazasini ishga tushirish
create_tables()

def login_required(f):
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

def superadmin_required(f):
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = User.get_by_id(session['user_id'])
        if not user.is_superadmin:
            flash('Sizda bu sahifaga kirish huquqi yo\'q', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    error_message = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        try:
            user = User.get(User.username == username)
            if user.check_password(password):
                session['user_id'] = user.id
                session['username'] = user.username
                session['is_superadmin'] = user.is_superadmin
                return redirect(url_for('dashboard'))
            else:
                error_message = 'Noto‘g‘ri login yoki parol'
                flash('Noto‘g‘ri login yoki parol', 'error')
        except User.DoesNotExist:
            error_message = 'Noto‘g‘ri login yoki parol'
            flash('Noto‘g‘ri login yoki parol', 'error')
    
    return render_template('login.html', error_message=error_message)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    # Statistika
    total_clients = Client.select().count()
    
    # Qarzdorlar (ijobiy balansga ega mijozlar - ular bizga qarzdor)
    debtors_count = Client.select().where(Client.balance > 0).count()
    debtors_sum = Client.select(fn.SUM(Client.balance)).where(Client.balance > 0).scalar() or 0
    
    # Kreditorlar (salbiy balansga ega mijozlar - biz ularga qarzdormiz)
    creditors_count = Client.select().where(Client.balance < 0).count()
    creditors_sum = abs(Client.select(fn.SUM(Client.balance)).where(Client.balance < 0).scalar() or 0)
    
    # Kirimlar (bugun)
    today = date.today()
    income_count = Transaction.select().where(
        Transaction.transaction_type == 'credit',
        Transaction.date == today
    ).count()
    income_sum = Transaction.select(fn.SUM(Transaction.amount)).where(
        Transaction.transaction_type == 'credit',
        Transaction.date == today
    ).scalar() or 0
    
    stats = {
        'income_count': income_count,
        'income_sum': income_sum,
        'debtors_count': debtors_count,
        'debtors_sum': debtors_sum,
        'creditors_count': creditors_count,
        'creditors_sum': creditors_sum
    }
    
    return render_template('dashboard.html', stats=stats)

@app.route('/clients')
@login_required
def clients():
    # Filtrlar
    name_filter = request.args.get('name', '')
    phone_filter = request.args.get('phone', '')
    date_filter = request.args.get('date', '')
    debtor_filter = request.args.get('debtor', '')
    
    query = Client.select()
    
    if name_filter:
        query = query.where(Client.name.contains(name_filter))
    if phone_filter:
        query = query.where(Client.phone.contains(phone_filter))
    if date_filter:
        query = query.where(Client.promised_payment_date == date_filter)
    if debtor_filter:
        if debtor_filter == 'debtor':
            query = query.where(Client.balance > 0)
        elif debtor_filter == 'creditor':
            query = query.where(Client.balance < 0)
    
    clients_list = list(query.order_by(Client.created_at.desc()))
    
    return render_template('clients.html', clients=clients_list)

@app.route('/clients/create', methods=['POST'])
@login_required
def create_client():
    try:
        print("Client creation request received")
        print("Form data:", request.form)
        print("Files:", request.files)
        
        # Ma'lumotlarni olish va tekshirish
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        comment = request.form.get('comment', '').strip()
        
        # Majburiy maydonlarni tekshirish
        if not name:
            flash('Mijoz nomi majburiy', 'error')
            return redirect(url_for('clients'))
        
        if not phone:
            flash('Telefon raqami majburiy', 'error')
            return redirect(url_for('clients'))
        
        # Telefon raqamini formatlash
        if phone and not phone.startswith('+'):
            phone = '+' + phone
        
        # Mijoz yaratish
        try:
            client = Client.create(
                name=name,
                phone=phone,
                comment=comment
            )
        except Exception as create_error:
            print(f"Client creation database error: {create_error}")
            flash('Mijoz yaratishda xatolik. Iltimos, qaytadan urinib ko\'ring', 'error')
            return redirect(url_for('clients'))
        
        # Rasm yuklash
        if 'photo' in request.files:
            photo_file = request.files['photo']
            if photo_file and photo_file.filename:
                try:
                    client.save_photo(photo_file)
                    client.save()
                except Exception as photo_error:
                    # Rasm yuklashda xatolik bo'lsa ham mijoz yaratilgan bo'ladi
                    print(f"Photo upload error: {photo_error}")
        
        flash(f'Mijoz "{name}" muvaffaqiyatli yaratildi', 'success')
        return redirect(url_for('clients'))
        
    except Exception as e:
        print(f"Client creation error: {e}")
        flash(f'Mijoz yaratishda xatolik: {str(e)}', 'error')
        return redirect(url_for('clients'))

@app.route('/clients/edit/<int:client_id>', methods=['GET', 'POST'])
@login_required
def edit_client(client_id):
    try:
        client = Client.get_by_id(client_id)
        
        if request.method == 'POST':
            try:
                name = request.form.get('name', '').strip()
                phone = request.form.get('phone', '').strip()
                comment = request.form.get('comment', '').strip()
                
                # Majburiy maydonlarni tekshirish
                if not name:
                    flash('Mijoz nomi majburiy', 'error')
                    return render_template('edit_client.html', client=client)
                
                if not phone:
                    flash('Telefon raqami majburiy', 'error')
                    return render_template('edit_client.html', client=client)
                
                # Telefon raqamini formatlash
                if phone and not phone.startswith('+'):
                    phone = '+' + phone
                
                client.name = name
                client.phone = phone
                client.comment = comment
                
                # Handle photo upload
                if 'photo' in request.files:
                    photo_file = request.files['photo']
                    if photo_file and photo_file.filename:
                        try:
                            client.save_photo(photo_file)
                        except Exception as photo_error:
                            print(f"Photo upload error: {photo_error}")
                            flash('Rasm yuklashda xatolik yuz berdi, lekin boshqa ma\'lumotlar saqlandi', 'warning')
                
                client.save()
                flash('Mijoz ma\'lumotlari muvaffaqiyatli yangilandi', 'success')
                return redirect(url_for('clients'))
                
            except Exception as e:
                print(f"Client edit error: {e}")
                flash('Mijoz ma\'lumotlarini yangilashda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
                return render_template('edit_client.html', client=client)
        
        return render_template('edit_client.html', client=client)
    
    except Client.DoesNotExist:
        flash('Mijoz topilmadi', 'error')
        return redirect(url_for('clients'))

@app.route('/clients/delete/<int:client_id>', methods=['POST'])
@login_required
def delete_client(client_id):
    try:
        client = Client.get_by_id(client_id)
        
        # Mijozning tranzaksiyalari bor-yo'qligini tekshiramiz
        if client.transactions.count() > 0:
            flash('Bu mijozda tranzaksiyalar mavjud. Avval ularni o\'chiring', 'error')
        else:
            try:
                # Delete photo file if exists
                if client.photo:
                    photo_path = os.path.join('static/uploads', client.photo)
                    if os.path.exists(photo_path):
                        os.remove(photo_path)
                
                client.delete_instance()
                flash('Mijoz muvaffaqiyatli o\'chirildi', 'success')
            except Exception as e:
                print(f"Client deletion error: {e}")
                flash('Mijozni o\'chirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
    
    except Client.DoesNotExist:
        flash('Mijoz topilmadi', 'error')
    except Exception as e:
        print(f"Client deletion error: {e}")
        flash('Mijozni o\'chirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
    
    return redirect(url_for('clients'))

@app.route('/debt')
@login_required
def debt():
    # Qidirish parametrlari
    search_query = request.args.get('search', '')
    
    if search_query:
        # Mijozlarni qidirish
        clients_list = list(Client.select().where(
            (Client.name.contains(search_query)) | 
            (Client.phone.contains(search_query))
        ).order_by(Client.name))
    else:
        clients_list = list(Client.select().order_by(Client.name))
    
    return render_template('debt.html', clients=clients_list, date=date, search_query=search_query)

@app.route('/debt/create', methods=['POST'])
@login_required
def create_debt():
    try:
        client_id = request.form.get('client_id')
        amount_str = request.form.get('amount')
        date_str = request.form.get('date')
        promised_date = request.form.get('promised_date')
        comment = request.form.get('comment', '')
        
        # Ma'lumotlarni tekshirish
        if not client_id:
            flash('Mijozni tanlang', 'error')
            return redirect(url_for('debt'))
        
        if not amount_str:
            flash('Qarz miqdorini kiriting', 'error')
            return redirect(url_for('debt'))
        
        if not date_str:
            flash('Qarz sanasini kiriting', 'error')
            return redirect(url_for('debt'))
        
        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                flash('Qarz miqdori 0 dan katta bo\'lishi kerak', 'error')
                return redirect(url_for('debt'))
        except (ValueError, TypeError):
            flash('Noto\'g\'ri qarz miqdori', 'error')
            return redirect(url_for('debt'))
        
        transaction_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Tranzaksiya yaratamiz
        Transaction.create(
            client_id=client_id,
            amount=amount,
            transaction_type='debit',  # mijozning qarzi
            date=transaction_date,
            comment=comment
        )
        
        # Mijozning balansini yangilaymiz
        client = Client.get_by_id(client_id)
        client.balance += amount
        
        # Va'da sanasini yangilaymiz
        if promised_date:
            client.promised_payment_date = datetime.strptime(promised_date, '%Y-%m-%d').date()
        
        client.save()
        
        flash(f'Qarz muvaffaqiyatli qo\'shildi. Miqdor: {amount:,.2f} so\'m', 'success')
        return redirect(url_for('debt'))
        
    except Exception as e:
        print(f"Debt creation error: {e}")
        flash('Qarz qo\'shishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
        return redirect(url_for('debt'))

@app.route('/finance')
@login_required
def finance():
    # Filtrlar
    client_filter = request.args.get('client', '')
    phone_filter = request.args.get('phone', '')
    type_filter = request.args.get('type', '')
    
    query = Transaction.select(Transaction, Client).join(Client)
    
    if client_filter:
        query = query.where(Client.name.contains(client_filter))
    if phone_filter:
        query = query.where(Client.phone.contains(phone_filter))
    if type_filter:
        query = query.where(Transaction.transaction_type == type_filter)
    
    transactions = list(query.order_by(Transaction.created_at.desc()))
    
    return render_template('finance.html', transactions=transactions)

@app.route('/settings')
@login_required
def settings():
    company_info = CompanyInfo.select().first()
    users = list(User.select().where(User.is_superadmin == False)) if session.get('is_superadmin') else []
    current_user = User.get_by_id(session['user_id'])
    
    return render_template('settings.html', 
                         company_info=company_info, 
                         users=users,
                         current_user=current_user,
                         is_superadmin=session.get('is_superadmin', False))

@app.route('/settings/company', methods=['POST'])
@login_required
def update_company():
    try:
        company_info = CompanyInfo.select().first()
        
        company_info.name = request.form.get('name', '').strip()
        company_info.phone = request.form.get('phone', '').strip()
        company_info.email = request.form.get('email', '').strip()
        company_info.address = request.form.get('address', '').strip()
        company_info.save()
        
        flash('Kompaniya ma\'lumotlari yangilandi', 'success')
        return redirect(url_for('settings'))
        
    except Exception as e:
        print(f"Company info update error: {e}")
        flash('Kompaniya ma\'lumotlarini yangilashda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
        return redirect(url_for('settings'))

@app.route('/settings/change_password', methods=['POST'])
@login_required
def change_password():
    try:
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        # Ma'lumotlarni tekshirish
        if not current_password:
            flash('Joriy parolni kiriting', 'error')
            return redirect(url_for('settings'))
        
        if not new_password:
            flash('Yangi parolni kiriting', 'error')
            return redirect(url_for('settings'))
        
        if not confirm_password:
            flash('Parolni tasdiqlang', 'error')
            return redirect(url_for('settings'))
        
        user = User.get_by_id(session['user_id'])
        
        if not user.check_password(current_password):
            flash('Joriy parol noto\'g\'ri', 'error')
            return redirect(url_for('settings'))
        
        if new_password != confirm_password:
            flash('Yangi parollar mos kelmadi', 'error')
            return redirect(url_for('settings'))
        
        if len(new_password) < 6:
            flash('Parol kamida 6 ta belgidan iborat bo\'lishi kerak', 'error')
            return redirect(url_for('settings'))
        
        user.change_password(new_password)
        flash('Parol muvaffaqiyatli o\'zgartirildi', 'success')
        return redirect(url_for('settings'))
        
    except Exception as e:
        print(f"Password change error: {e}")
        flash('Parolni o\'zgartirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
        return redirect(url_for('settings'))

@app.route('/settings/create_admin', methods=['POST'])
@superadmin_required
def create_admin():
    try:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        # Ma'lumotlarni tekshirish
        if not username:
            flash('Foydalanuvchi nomini kiriting', 'error')
            return redirect(url_for('settings'))
        
        if not password:
            flash('Parolni kiriting', 'error')
            return redirect(url_for('settings'))
        
        if len(password) < 6:
            flash('Parol kamida 6 ta belgidan iborat bo\'lishi kerak', 'error')
            return redirect(url_for('settings'))
        
        try:
            User.create_user(username=username, password=password, is_superadmin=False)
            flash(f'Administrator "{username}" muvaffaqiyatli yaratildi', 'success')
        except IntegrityError:
            flash('Bunday foydalanuvchi nomi allaqachon mavjud', 'error')
        except Exception as e:
            print(f"Admin creation error: {e}")
            flash('Administrator yaratishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
        
        return redirect(url_for('settings'))
        
    except Exception as e:
        print(f"Admin creation error: {e}")
        flash('Administrator yaratishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
        return redirect(url_for('settings'))

@app.route('/settings/delete_admin/<int:user_id>', methods=['POST'])
@superadmin_required
def delete_admin(user_id):
    try:
        user = User.get_by_id(user_id)
        if user.is_superadmin:
            flash('Super adminni o\'chirish mumkin emas', 'error')
        else:
            try:
                username = user.username
                user.delete_instance()
                flash(f'Administrator "{username}" muvaffaqiyatli o\'chirildi', 'success')
            except Exception as e:
                print(f"Admin deletion error: {e}")
                flash('Administratorni o\'chirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
    except User.DoesNotExist:
        flash('Foydalanuvchi topilmadi', 'error')
    except Exception as e:
        print(f"Admin deletion error: {e}")
        flash('Administratorni o\'chirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
    
    return redirect(url_for('settings'))

@app.route('/payment/<int:client_id>', methods=['GET', 'POST'])
@login_required
def payment(client_id):
    try:
        client = Client.get_by_id(client_id)
        
        if request.method == 'POST':
            try:
                payment_type = request.form.get('payment_type')
                date_str = request.form.get('date')
                comment = request.form.get('comment', '')
                
                # Ma'lumotlarni tekshirish
                if not payment_type:
                    flash('To\'lov turini tanlang', 'error')
                    return render_template('payment.html', client=client, date=date)
                
                if not date_str:
                    flash('To\'lov sanasini kiriting', 'error')
                    return render_template('payment.html', client=client, date=date)
                
                transaction_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                # To'lov miqdorini hisoblash
                if payment_type == 'full':
                    payment_amount = client.balance
                elif payment_type == 'partial':
                    try:
                        amount = Decimal(request.form.get('amount', '0'))
                        if amount <= 0:
                            flash('To\'lov miqdori 0 dan katta bo\'lishi kerak', 'error')
                            return render_template('payment.html', client=client, date=date)
                        
                        if amount > client.balance:
                            flash('To\'lov miqdori qarz miqdoridan katta bo\'lishi mumkin emas', 'error')
                            return render_template('payment.html', client=client, date=date)
                        
                        payment_amount = amount
                    except (ValueError, TypeError):
                        flash('Noto\'g\'ri to\'lov miqdori', 'error')
                        return render_template('payment.html', client=client, date=date)
                else:
                    flash('Noto\'g\'ri to\'lov turi', 'error')
                    return render_template('payment.html', client=client, date=date)
                
                # Tranzaksiya yaratamiz
                Transaction.create(
                    client_id=client_id,
                    amount=payment_amount,
                    transaction_type='credit',  # to'lov - kompaniyaning qarzi
                    date=transaction_date,
                    comment=comment
                )
                
                # Mijozning balansini yangilaymiz
                client.balance -= payment_amount
                client.save()
                
                flash(f'To\'lov muvaffaqiyatli amalga oshirildi. Miqdor: {payment_amount:,.2f} so\'m', 'success')
                return redirect(url_for('clients'))
                
            except Exception as e:
                print(f"Payment error: {e}")
                flash('To\'lov amalga oshirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
                return render_template('payment.html', client=client, date=date)
        
        return render_template('payment.html', client=client, date=date)
    
    except Client.DoesNotExist:
        flash('Mijoz topilmadi', 'error')
        return redirect(url_for('clients'))

@app.route('/payment/history/<int:client_id>')
@login_required
def payment_history(client_id):
    try:
        client = Client.get_by_id(client_id)
        transactions = list(Transaction.select().where(Transaction.client_id == client_id).order_by(Transaction.date.desc()))
        
        return render_template('payment_history.html', client=client, transactions=transactions)
    
    except Client.DoesNotExist:
        flash('Mijoz topilmadi', 'error')
        return redirect(url_for('clients'))

@app.route('/transaction/edit/<int:transaction_id>', methods=['GET', 'POST'])
@login_required
def edit_transaction(transaction_id):
    try:
        transaction = Transaction.get_by_id(transaction_id)
        
        if request.method == 'POST':
            try:
                # Eski balansni qaytarib olamiz
                client = transaction.client
                if transaction.transaction_type == 'debit':
                    client.balance -= transaction.amount
                else:
                    client.balance += transaction.amount
                
                # Yangi ma'lumotlarni olamiz
                amount_str = request.form.get('amount')
                transaction_type = request.form.get('transaction_type')
                date_str = request.form.get('date')
                comment = request.form.get('comment', '')
                
                # Ma'lumotlarni tekshirish
                if not amount_str:
                    flash('Tranzaksiya miqdorini kiriting', 'error')
                    return render_template('edit_transaction.html', transaction=transaction)
                
                if not transaction_type:
                    flash('Tranzaksiya turini tanlang', 'error')
                    return render_template('edit_transaction.html', transaction=transaction)
                
                if not date_str:
                    flash('Tranzaksiya sanasini kiriting', 'error')
                    return render_template('edit_transaction.html', transaction=transaction)
                
                try:
                    amount = Decimal(amount_str)
                    if amount <= 0:
                        flash('Tranzaksiya miqdori 0 dan katta bo\'lishi kerak', 'error')
                        return render_template('edit_transaction.html', transaction=transaction)
                except (ValueError, TypeError):
                    flash('Noto\'g\'ri tranzaksiya miqdori', 'error')
                    return render_template('edit_transaction.html', transaction=transaction)
                
                transaction_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                # Tranzaksiyani yangilaymiz
                transaction.amount = amount
                transaction.transaction_type = transaction_type
                transaction.date = transaction_date
                transaction.comment = comment
                transaction.save()
                
                # Yangi balansni hisoblaymiz
                if transaction_type == 'debit':
                    client.balance += amount
                else:
                    client.balance -= amount
                
                client.save()
                
                flash('Tranzaksiya muvaffaqiyatli yangilandi', 'success')
                return redirect(url_for('finance'))
                
            except Exception as e:
                print(f"Transaction edit error: {e}")
                flash('Tranzaksiyani yangilashda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
                return render_template('edit_transaction.html', transaction=transaction)
        
        return render_template('edit_transaction.html', transaction=transaction)
    
    except Transaction.DoesNotExist:
        flash('Tranzaksiya topilmadi', 'error')
        return redirect(url_for('finance'))

@app.route('/transaction/delete/<int:transaction_id>', methods=['POST'])
@login_required
def delete_transaction(transaction_id):
    try:
        transaction = Transaction.get_by_id(transaction_id)
        client = transaction.client
        
        # Balansni qaytarib olamiz
        if transaction.transaction_type == 'debit':
            client.balance -= transaction.amount
        else:
            client.balance += transaction.amount
        
        client.save()
        
        # Tranzaksiyani o'chiramiz
        transaction.delete_instance()
        
        flash('Tranzaksiya muvaffaqiyatli o\'chirildi', 'success')
    
    except Transaction.DoesNotExist:
        flash('Tranzaksiya topilmadi', 'error')
    except Exception as e:
        print(f"Transaction deletion error: {e}")
        flash('Tranzaksiyani o\'chirishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
    
    return redirect(url_for('finance'))

@app.route('/sms')
@login_required
def sms():
    return render_template('sms.html')

def generate_excel_report(report_type='all'):
    """Excel hisobot yaratish"""
    wb = Workbook()
    
    # Default sheet o'chirish
    wb.remove(wb.active)
    
    if report_type == 'all' or report_type == 'clients':
        # Mijozlar hisoboti
        ws_clients = wb.create_sheet("Mijozlar")
        
        # Sarlavhalar
        headers = ['№', 'Mijoz nomi', 'Telefon', 'Balans', 'Va\'da sanasi', 'Izoh', 'Yaratilgan sana']
        for col, header in enumerate(headers, 1):
            cell = ws_clients.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Ma'lumotlar
        clients = Client.select().order_by(Client.name)
        for row, client in enumerate(clients, 2):
            ws_clients.cell(row=row, column=1, value=row-1)
            ws_clients.cell(row=row, column=2, value=client.name)
            ws_clients.cell(row=row, column=3, value=client.phone)
            ws_clients.cell(row=row, column=4, value=float(client.balance))
            ws_clients.cell(row=row, column=5, value=client.promised_payment_date.strftime('%d.%m.%Y') if client.promised_payment_date else '')
            ws_clients.cell(row=row, column=6, value=client.comment or '')
            ws_clients.cell(row=row, column=7, value=client.created_at.strftime('%d.%m.%Y'))
        
        # Ustun kengliklarini sozlash
        for col in range(1, len(headers) + 1):
            ws_clients.column_dimensions[get_column_letter(col)].width = 15
    
    if report_type == 'all' or report_type == 'transactions':
        # Tranzaksiyalar hisoboti
        ws_transactions = wb.create_sheet("Tranzaksiyalar")
        
        # Sarlavhalar
        headers = ['№', 'Mijoz', 'Telefon', 'Miqdor', 'Turi', 'Sana', 'Izoh']
        for col, header in enumerate(headers, 1):
            cell = ws_transactions.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Ma'lumotlar
        transactions = Transaction.select(Transaction, Client).join(Client).order_by(Transaction.date.desc())
        for row, transaction in enumerate(transactions, 2):
            ws_transactions.cell(row=row, column=1, value=row-1)
            ws_transactions.cell(row=row, column=2, value=transaction.client.name)
            ws_transactions.cell(row=row, column=3, value=transaction.client.phone)
            ws_transactions.cell(row=row, column=4, value=float(transaction.amount))
            ws_transactions.cell(row=row, column=5, value='Qarz' if transaction.transaction_type == 'debit' else 'To\'lov')
            ws_transactions.cell(row=row, column=6, value=transaction.date.strftime('%d.%m.%Y'))
            ws_transactions.cell(row=row, column=7, value=transaction.comment or '')
        
        # Ustun kengliklarini sozlash
        for col in range(1, len(headers) + 1):
            ws_transactions.column_dimensions[get_column_letter(col)].width = 15
    
    if report_type == 'all' or report_type == 'summary':
        # Umumiy hisobot
        ws_summary = wb.create_sheet("Umumiy hisobot")
        
        # Statistika
        total_clients = Client.select().count()
        debtors_count = Client.select().where(Client.balance > 0).count()
        debtors_sum = Client.select(fn.SUM(Client.balance)).where(Client.balance > 0).scalar() or 0
        creditors_count = Client.select().where(Client.balance < 0).count()
        creditors_sum = abs(Client.select(fn.SUM(Client.balance)).where(Client.balance < 0).scalar() or 0)
        
        # Bugungi kirimlar
        today = date.today()
        income_count = Transaction.select().where(
            Transaction.transaction_type == 'credit',
            Transaction.date == today
        ).count()
        income_sum = Transaction.select(fn.SUM(Transaction.amount)).where(
            Transaction.transaction_type == 'credit',
            Transaction.date == today
        ).scalar() or 0
        
        # Ma'lumotlar
        summary_data = [
            ['Statistika', 'Qiymat'],
            ['Jami mijozlar', total_clients],
            ['Qarzdorlar soni', debtors_count],
            ['Qarzdorlik summasi', f"{float(debtors_sum):,.2f} so'm"],
            ['Kreditorlar soni', creditors_count],
            ['Kreditorlik summasi', f"{float(creditors_sum):,.2f} so'm"],
            ['Bugungi kirimlar soni', income_count],
            ['Bugungi kirimlar summasi', f"{float(income_sum):,.2f} so'm"],
            ['', ''],
            ['Hisobot sanasi', datetime.now().strftime('%d.%m.%Y %H:%M')]
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row, column=1, value=label)
            ws_summary.cell(row=row, column=2, value=value)
            
            # Sarlavha stillarini qo'shish
            if row == 1:
                for col in range(1, 3):
                    cell = ws_summary.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")
        
        # Ustun kengliklarini sozlash
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
    
    # Excel faylini saqlash
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html')

@app.route('/reports/download/<report_type>')
@login_required
def download_report(report_type):
    try:
        if report_type not in ['all', 'clients', 'transactions', 'summary']:
            flash('Noto\'g\'ri hisobot turi', 'error')
            return redirect(url_for('reports'))
        
        # Excel faylini yaratish
        excel_file = generate_excel_report(report_type)
        
        # Fayl nomini belgilash
        report_names = {
            'all': 'Umumiy_hisobot',
            'clients': 'Mijozlar_hisoboti',
            'transactions': 'Tranzaksiyalar_hisoboti',
            'summary': 'Statistika_hisoboti'
        }
        
        filename = f"{report_names[report_type]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Report generation error: {e}")
        flash('Hisobot yaratishda xatolik yuz berdi. Iltimos, qaytadan urinib ko\'ring', 'error')
        return redirect(url_for('reports'))

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)